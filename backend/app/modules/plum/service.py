"""
PLUM service layer (business logic).

Phase 5: Part CRUD, revision FSM, search/filter, auto-generated part numbers,
and revision label generation.

Phase 6: BOM CRUD (Draft-only, cycle detection), BOM traversal (tree/flat/where-used),
AVL link + price-break CRUD (vendor validation), effective-cost resolution chain (D-07),
margin computation (D-09), as-released cost snapshot hook (D-14), BOM copy-forward
on new revision (D-01), system currency helper (_get_system_currency).

Part number generation (D-06):
  Numbers follow the series "P00001", "P00002", … using a DB MAX query.
  The unique DB constraint on plum_part.part_number is the real guard against
  duplicates (not application-level locking). On an IntegrityError collision,
  the function retries once with a freshly generated number (RESEARCH Pattern 7).

Revision labels (D-04/D-05):
  ASME scheme: A, B, C, ... (skipping I, O, Q, S, X, Z per ASME Y14.35).
  SemVer scheme: 0.1.0 → 1.0.0 on release; new Draft minor-bumps (1.0.0 → 1.1.0).
  Read from the global plum.revision_scheme setting; defaults to "asme".

Revision FSM (D-07):
  VALID_TRANSITIONS maps each status to the list of allowed target statuses.
  advance_revision_status validates and applies the transition; on →released,
  the prior released revision is auto-obsoleted in the same transaction (D-08).
  Pitfall 3: flush() between the obsolete update and the release update to
  avoid the partial unique index seeing two released rows simultaneously.

Soft-delete (D-11):
  Parts are never hard-deleted. Setting active=False hides from default lists.

Immutability (D-07):
  Released revisions are frozen. update_part raises 422 if revision-controlled
  fields are changed while the current revision is "released".
  BOM lines can only be edited on Draft revisions (T-06-05).

Server-side search (D-15/PLUM-02):
  list_parts uses parameterized SQLAlchemy .ilike() — never raw-SQL
  interpolation — to satisfy T-05-06 (search threat mitigation).

BOM cycle detection (D-05):
  _would_create_cycle uses BFS from candidate_child downward through the BOM
  graph. If the parent_part_id appears in candidate_child's descendants, adding
  would create a cycle → rejected with 422.

Effective-cost chain (D-07):
  compute_effective_cost resolves in priority order:
    1. selected vendor price-break (bounds-checked against actual list length)
    2. manual material_cost
    3. BOM roll-up Σ(child effective_cost × qty)
    4. None / "uncosted"
  Manual wins over roll-up even on assemblies (purchased sub-assembly support).

No ORM relationship access anywhere in this module (MissingGreenlet pitfall).
All queries use explicit select(...) calls.
"""
from __future__ import annotations

import uuid
from datetime import datetime, timezone
from decimal import Decimal
from typing import TYPE_CHECKING

from fastapi import HTTPException, status
from sqlalchemy import Numeric, cast, func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession

if TYPE_CHECKING:
    from app.modules.plum.models import PlumBomItem, PlumPart, PlumPartRevision, PlumPartTag
    from app.modules.plum.schemas import (
        AvlLinkCreate,
        AvlLinkUpdate,
        BomItemCreate,
        BomItemUpdate,
        CostUpdate,
        PartCreate,
        PartUpdate,
        PriceBreakCreate,
        RevisionCreate,
    )


# ---------------------------------------------------------------------------
# ASME revision letter constants (D-04)
# ---------------------------------------------------------------------------

ASME_SKIP: set[str] = {"I", "O", "Q", "S", "X", "Z"}
ASME_LETTERS: list[str] = [c for c in "ABCDEFGHIJKLMNOPQRSTUVWXYZ" if c not in ASME_SKIP]

# ---------------------------------------------------------------------------
# Revision FSM transition table (D-07)
# ---------------------------------------------------------------------------

VALID_TRANSITIONS: dict[str, list[str]] = {
    "draft":     ["in_review"],
    "in_review": ["released", "draft"],
    "released":  ["obsolete"],  # internal only — triggered by supersede (D-08)
    "obsolete":  [],            # terminal — no outgoing transitions
}


# ---------------------------------------------------------------------------
# Part number generation (D-06)
# ---------------------------------------------------------------------------


async def generate_part_number(db: AsyncSession) -> str:
    """
    Generate the next part number in the P##### series.

    Finds the current highest *numeric* suffix among strictly-numeric P-series
    part numbers (matching ``^P[0-9]+$``) by casting the digits after "P" to a
    number and ordering numerically, then returns the next value zero-padded to
    5 digits. Returns "P00001" when no P-series part numbers exist yet.

    The regex filter MUST precede the cast: a bare cast over ``LIKE 'P%'`` would
    throw on non-numeric part numbers such as "P-DUPE-01" (PLUM-01 defect). The
    previous implementation used lexicographic MAX(part_number), which returned a
    smaller successor once the suffix crossed a digit-width boundary (e.g.
    "P99999" sorted above "P100000"), producing duplicate part numbers.

    The cast target is Numeric, not Integer: part_number is String(50) with no
    format constraint, so a caller may legally create "P9999999999", whose suffix
    matches the regex but exceeds int4 — an Integer cast would raise "value out of
    range for type integer" here and 500 every subsequent auto-numbered create
    until the row was deleted by hand. Numeric cannot overflow for any 50-char
    digit string.

    The DB unique constraint on plum_part.part_number is the authoritative guard;
    this function is a best-effort generator. The caller must handle
    IntegrityError on collision (RESEARCH Pattern 7).
    """
    from app.modules.plum.models import PlumPart

    result = await db.execute(
        select(PlumPart.part_number)
        .where(PlumPart.part_number.op("~")(r"^P[0-9]+$"))
        .order_by(cast(func.substring(PlumPart.part_number, 2), Numeric).desc())
        .limit(1)
    )
    max_pn: str | None = result.scalar()

    if max_pn is None:
        return "P00001"

    # Parse the numeric suffix after "P"
    try:
        suffix = int(max_pn[1:])
    except (IndexError, ValueError):
        suffix = 0

    return f"P{suffix + 1:05d}"


# ---------------------------------------------------------------------------
# Revision scheme helpers (D-04/D-05)
# ---------------------------------------------------------------------------


async def _get_revision_scheme(db: AsyncSession) -> str:
    """
    Read the plum.revision_scheme setting from the global settings table.

    Returns "asme" if the setting is absent (safe default — ASME is the
    initial seed value in seed_plum_data).
    """
    from app.core.settings_model import Setting

    result = await db.execute(
        select(Setting.value).where(Setting.key == "plum.revision_scheme")
    )
    value: str | None = result.scalar()
    return value or "asme"


async def _get_system_currency(db: AsyncSession) -> str:
    """
    Read the locale.currency setting from the global settings table.

    Returns "USD" if the setting is absent. Mirrors _get_revision_scheme idiom.
    D-10: single system currency seeded as "USD" in Phase-3 settings seed.
    """
    from app.core.settings_model import Setting

    result = await db.execute(
        select(Setting.value).where(Setting.key == "locale.currency")
    )
    value: str | None = result.scalar()
    return value or "USD"


def _first_revision_label(scheme: str) -> str:
    """
    Return the label for the first revision under the given scheme (D-04).

    ASME: "A" (first letter in the valid alphabet).
    SemVer: "0.1.0" (pre-release initial draft).
    """
    if scheme == "semver":
        return "0.1.0"
    # Default: ASME
    return ASME_LETTERS[0]  # "A"


def _next_draft_label(scheme: str, source_label: str) -> str:
    """
    Return the label for a new Draft revision copied from source_label (D-05).

    ASME: advance to the next letter (skipping ASME_SKIP letters).
      After the last single letter "Y", wraps to "AA", "AB", ...
    SemVer: minor-bump the source label.
      - If source is a release (e.g. "1.0.0"), next draft minor-bumps: "1.1.0".
      - If source is already a draft (e.g. "0.1.0"), minor-bumps: "0.2.0".
      - Patch is reset to 0 on each new draft.
    """
    if scheme == "semver":
        return _semver_minor_bump(source_label)
    # ASME scheme
    return _asme_next_letter(source_label)


def _release_label(scheme: str, current_label: str) -> str:
    """
    Return the label to apply on release of a draft (D-05).

    ASME: label is set at creation; no change on release — return unchanged.
    SemVer: major-bump (zeros rest): "0.1.0" → "1.0.0", "1.1.0" → "2.0.0".
    """
    if scheme == "semver":
        return _semver_major_bump(current_label)
    # ASME: label is set at creation time; unchanged on release
    return current_label


def _asme_next_letter(current_label: str) -> str:
    """Advance to the next ASME letter, wrapping to double-letter after 'Y'."""
    upper = current_label.upper()
    # Simple single-letter case
    if len(upper) == 1:
        if upper in ASME_LETTERS:
            idx = ASME_LETTERS.index(upper)
            if idx + 1 < len(ASME_LETTERS):
                return ASME_LETTERS[idx + 1]
            # Exhausted single letters — wrap to "AA"
            return "AA"
    # Multi-letter — increment last character
    if len(upper) == 2:
        prefix = upper[0]
        last = upper[1]
        if last in ASME_LETTERS:
            idx = ASME_LETTERS.index(last)
            if idx + 1 < len(ASME_LETTERS):
                return prefix + ASME_LETTERS[idx + 1]
            # Advance prefix letter
            if prefix in ASME_LETTERS:
                pidx = ASME_LETTERS.index(prefix)
                if pidx + 1 < len(ASME_LETTERS):
                    return ASME_LETTERS[pidx + 1] + ASME_LETTERS[0]
    # Fallback: append a marker to signal exhaustion (extremely unlikely in v1)
    return upper + ASME_LETTERS[0]


def _semver_minor_bump(label: str) -> str:
    """Minor-bump a SemVer label (zeroes patch)."""
    try:
        parts = label.split(".")
        major = int(parts[0])
        minor = int(parts[1])
        return f"{major}.{minor + 1}.0"
    except (IndexError, ValueError):
        return "0.2.0"


def _semver_major_bump(label: str) -> str:
    """Major-bump a SemVer label (zeroes minor and patch)."""
    try:
        parts = label.split(".")
        major = int(parts[0])
        return f"{major + 1}.0.0"
    except (IndexError, ValueError):
        return "1.0.0"


# ---------------------------------------------------------------------------
# Part CRUD
# ---------------------------------------------------------------------------


async def create_part(db: AsyncSession, data: "PartCreate") -> "PlumPart":
    """
    Insert a new part + auto-create its first Draft revision (D-03).

    If data.part_number is not supplied, auto-generates one via generate_part_number.
    On a unique-constraint IntegrityError:
      - If the caller explicitly supplied a part_number → 409 Conflict.
      - If the part_number was auto-generated (race condition) → retry ONCE.

    First revision is always created in "draft" status with revision_number=1
    and a label derived from the plum.revision_scheme setting. Description and
    other revision-controlled fields are seeded from data (D-03).

    Returns the refreshed PlumPart ORM instance (no revision attached — caller
    must call get_part_with_revisions for the full object, or use the ORM instance
    for list serialization with current_revision_* fields populated separately).
    """
    import sqlalchemy.exc

    from app.modules.plum.models import PlumPart, PlumPartRevision, PlumPartTag

    user_supplied_pn = bool(data.part_number)
    part_number = data.part_number or await generate_part_number(db)

    # Read the revision scheme before any write (avoid extra round-trip after flush)
    scheme = await _get_revision_scheme(db)
    first_label = _first_revision_label(scheme)

    part = PlumPart(
        id=str(uuid.uuid4()),
        part_number=part_number,
        active=True,
    )
    db.add(part)

    try:
        await db.flush()
    except sqlalchemy.exc.IntegrityError:
        await db.rollback()

        if user_supplied_pn:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail=f"Part number '{part_number}' already exists.",
            )

        # Auto-generated collision — retry once with a fresh number
        part_number = await generate_part_number(db)
        part = PlumPart(
            id=str(uuid.uuid4()),
            part_number=part_number,
            active=True,
        )
        db.add(part)
        await db.flush()

    # Auto-create first Draft revision (D-03)
    revision = PlumPartRevision(
        id=str(uuid.uuid4()),
        part_id=part.id,
        revision_number=1,
        revision_label=first_label,
        status="draft",
        description=data.description,
        category=data.category,
        unit_of_measure=data.unit_of_measure,
        notes=data.notes,
        reason_for_revision=data.reason_for_revision,
    )
    db.add(revision)

    # Insert classification tag join rows (D-12)
    for tag_id in data.tag_ids:
        db.add(PlumPartTag(part_id=part.id, tag_id=tag_id))

    await db.commit()
    await db.refresh(part)
    return part


async def list_parts(
    db: AsyncSession,
    q: str | None = None,
    status_filter: str | None = None,
    include_archived: bool = False,
) -> list[dict]:
    """
    Return parts matching the given filters, with current-revision summary fields.

    Args:
        q: Case-insensitive substring search across part_number OR current
           revision description. Uses parameterized .ilike() (T-05-06).
        status_filter: Filter parts whose current (latest by revision_number)
           revision has this status. Implemented as a MAX-based correlated
           subquery (RESEARCH Pattern 4 / PLUM-02).
        include_archived: When False (default), excludes active=False parts.

    Returns a list of dicts shaped for PartRead serialization, including
    current_revision_label, current_revision_status, and tag name list.
    """
    from app.modules.plum.models import PlumClassificationTag, PlumPart, PlumPartRevision, PlumPartTag

    # Subquery: max revision_number per part_id (latest revision)
    latest_rev_sq = (
        select(
            PlumPartRevision.part_id.label("part_id"),
            func.max(PlumPartRevision.revision_number).label("max_rev_num"),
        )
        .group_by(PlumPartRevision.part_id)
        .subquery()
    )

    # Join PlumPart → latest revision
    stmt = (
        select(PlumPart, PlumPartRevision)
        .join(
            latest_rev_sq,
            latest_rev_sq.c.part_id == PlumPart.id,
            isouter=True,
        )
        .join(
            PlumPartRevision,
            (PlumPartRevision.part_id == PlumPart.id)
            & (PlumPartRevision.revision_number == latest_rev_sq.c.max_rev_num),
            isouter=True,
        )
    )

    if not include_archived:
        stmt = stmt.where(PlumPart.active == True)  # noqa: E712

    if q:
        like = f"%{q}%"
        stmt = stmt.where(
            or_(
                PlumPart.part_number.ilike(like),
                PlumPartRevision.description.ilike(like),
            )
        )

    if status_filter:
        stmt = stmt.where(PlumPartRevision.status == status_filter)

    stmt = stmt.order_by(PlumPart.part_number)

    result = await db.execute(stmt)
    rows = result.all()

    # For each part, fetch tag names
    part_ids = [row[0].id for row in rows]
    tag_map: dict[str, list[str]] = {pid: [] for pid in part_ids}

    if part_ids:
        tag_stmt = (
            select(PlumPartTag.part_id, PlumClassificationTag.name)
            .join(PlumClassificationTag, PlumClassificationTag.id == PlumPartTag.tag_id)
            .where(PlumPartTag.part_id.in_(part_ids))
        )
        tag_result = await db.execute(tag_stmt)
        for part_id, tag_name in tag_result.all():
            tag_map[part_id].append(tag_name)

    # Shape results for PartRead serialization (current_revision_* are not ORM cols)
    parts_out = []
    for part, revision in rows:
        parts_out.append(
            {
                "id": part.id,
                "part_number": part.part_number,
                "active": part.active,
                "created_at": part.created_at,
                "updated_at": part.updated_at,
                "current_revision_label": revision.revision_label if revision else None,
                "current_revision_status": revision.status if revision else None,
                "tags": tag_map.get(part.id, []),
            }
        )

    return parts_out


async def get_part(db: AsyncSession, part_id: str) -> "PlumPart":
    """
    Load a part by id.

    Raises HTTP 404 if no part with the given id exists.
    """
    from app.modules.plum.models import PlumPart

    result = await db.execute(select(PlumPart).where(PlumPart.id == part_id))
    part = result.scalars().first()

    if part is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Part {part_id} not found",
        )

    return part


async def get_part_with_revisions(db: AsyncSession, part_id: str) -> dict:
    """
    Return the part plus all its revisions ordered newest-first (D-14).

    Shaped for PartDetailRead serialization. Raises 404 if part not found.
    """
    from app.modules.plum.models import PlumClassificationTag, PlumPart, PlumPartRevision, PlumPartTag

    # Load the part (raises 404 if missing)
    part = await get_part(db, part_id)

    # Load all revisions for this part, newest-first (highest revision_number first)
    rev_result = await db.execute(
        select(PlumPartRevision)
        .where(PlumPartRevision.part_id == part_id)
        .order_by(PlumPartRevision.revision_number.desc())
    )
    revisions = list(rev_result.scalars().all())

    # Load tag names for this part
    tag_stmt = (
        select(PlumClassificationTag.name)
        .join(PlumPartTag, PlumPartTag.tag_id == PlumClassificationTag.id)
        .where(PlumPartTag.part_id == part_id)
    )
    tag_result = await db.execute(tag_stmt)
    tag_names = [row[0] for row in tag_result.all()]

    return {
        "id": part.id,
        "part_number": part.part_number,
        "active": part.active,
        "created_at": part.created_at,
        "updated_at": part.updated_at,
        "tags": tag_names,
        "revisions": revisions,
    }


async def update_part(
    db: AsyncSession,
    part_id: str,
    data: "PartUpdate",
) -> dict:
    """
    Apply a partial update to a part (PATCH semantics).

    Part-level fields (part_number, active, tag_ids) are applied to PlumPart /
    PlumPartTag. Revision-controlled fields (description, category, unit_of_measure,
    notes) are applied to the current Draft revision.

    Raises HTTP 422 if revision-controlled fields are supplied and the current
    revision status is "released" (D-07 immutability, T-05-07).
    Raises HTTP 404 if the part does not exist.

    Returns a dict shaped for PartRead (with current_revision_* fields populated).
    """
    from app.modules.plum.models import PlumClassificationTag, PlumPart, PlumPartRevision, PlumPartTag

    part = await get_part(db, part_id)
    update_data = data.model_dump(exclude_unset=True)

    # Separate part-level from revision-controlled fields
    revision_fields = {"description", "category", "unit_of_measure", "notes"}
    part_level_fields = {"part_number", "active"}
    tag_field = "tag_ids"

    rev_updates = {k: v for k, v in update_data.items() if k in revision_fields}

    if rev_updates:
        # Load current (latest) revision to check immutability
        rev_result = await db.execute(
            select(PlumPartRevision)
            .where(PlumPartRevision.part_id == part_id)
            .order_by(PlumPartRevision.revision_number.desc())
        )
        current_rev = rev_result.scalars().first()

        if current_rev and current_rev.status == "released":
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="Cannot edit revision-controlled fields on a Released revision (D-07).",
            )

        # Apply revision-controlled updates to current Draft revision
        if current_rev:
            for field, value in rev_updates.items():
                setattr(current_rev, field, value)

    # Apply part-level fields
    for field in part_level_fields:
        if field in update_data:
            setattr(part, field, update_data[field])

    # Update tags if provided: replace join-table rows
    if tag_field in update_data:
        # Delete existing tag rows for this part
        existing_tags_result = await db.execute(
            select(PlumPartTag).where(PlumPartTag.part_id == part_id)
        )
        for tag_row in existing_tags_result.scalars().all():
            await db.delete(tag_row)
        # Insert new tag rows
        for tag_id in update_data[tag_field]:
            db.add(PlumPartTag(part_id=part_id, tag_id=tag_id))

    await db.commit()
    await db.refresh(part)

    # Fetch current revision info for PartRead response
    rev_result2 = await db.execute(
        select(PlumPartRevision)
        .where(PlumPartRevision.part_id == part_id)
        .order_by(PlumPartRevision.revision_number.desc())
    )
    current_rev2 = rev_result2.scalars().first()

    # Fetch tag names
    tag_stmt = (
        select(PlumClassificationTag.name)
        .join(PlumPartTag, PlumPartTag.tag_id == PlumClassificationTag.id)
        .where(PlumPartTag.part_id == part_id)
    )
    tag_result = await db.execute(tag_stmt)
    tag_names = [row[0] for row in tag_result.all()]

    return {
        "id": part.id,
        "part_number": part.part_number,
        "active": part.active,
        "created_at": part.created_at,
        "updated_at": part.updated_at,
        "current_revision_label": current_rev2.revision_label if current_rev2 else None,
        "current_revision_status": current_rev2.status if current_rev2 else None,
        "tags": tag_names,
    }


# ---------------------------------------------------------------------------
# Revision service
# ---------------------------------------------------------------------------


async def get_revision(db: AsyncSession, revision_id: str) -> "PlumPartRevision":
    """
    Load a revision by id.

    Raises HTTP 404 if no revision with the given id exists.
    """
    from app.modules.plum.models import PlumPartRevision

    result = await db.execute(
        select(PlumPartRevision).where(PlumPartRevision.id == revision_id)
    )
    revision = result.scalars().first()

    if revision is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Revision {revision_id} not found",
        )

    return revision


async def get_released_revision(
    db: AsyncSession, part_id: str
) -> "PlumPartRevision | None":
    """
    Return the currently Released revision for a part, or None if none exists.

    Used in the supersede flow (D-08): when a new revision is released, this
    function finds the prior released revision to obsolete.
    """
    from app.modules.plum.models import PlumPartRevision

    result = await db.execute(
        select(PlumPartRevision).where(
            PlumPartRevision.part_id == part_id,
            PlumPartRevision.status == "released",
        )
    )
    return result.scalars().first()


async def create_revision(
    db: AsyncSession,
    part_id: str,
    data: "RevisionCreate",
    actor_id: str,
) -> "PlumPartRevision":
    """
    Create a new Draft revision for a part (D-03 copy-forward).

    Source resolution order (D-03):
      1. data.source_revision_id if explicitly provided.
      2. Latest Released revision for the part.
      3. Latest overall revision (by revision_number) if no released one exists.

    Copies description, category, unit_of_measure, notes from the source.
    If RevisionCreate provides overrides, they take precedence over the copy.

    revision_number = MAX(existing revision_number for this part) + 1.
    revision_label  = _next_draft_label(scheme, source_label).
    status          = "draft".

    Writes a "revision.created" audit event.
    Returns the new PlumPartRevision ORM instance.
    """
    from app.modules.auth.service import write_audit
    from app.modules.plum.models import PlumPartRevision

    # Verify part exists (raises 404 if not)
    await get_part(db, part_id)

    scheme = await _get_revision_scheme(db)

    # Resolve source revision
    if data.source_revision_id:
        source = await get_revision(db, data.source_revision_id)
        if source.part_id != part_id:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Revision {data.source_revision_id} does not belong to part {part_id}",
            )
    else:
        # Default: latest Released, fallback to latest overall
        source = await get_released_revision(db, part_id)
        if source is None:
            # No released revision — use latest by revision_number
            latest_result = await db.execute(
                select(PlumPartRevision)
                .where(PlumPartRevision.part_id == part_id)
                .order_by(PlumPartRevision.revision_number.desc())
            )
            source = latest_result.scalars().first()

    # Determine next revision_number
    max_num_result = await db.execute(
        select(func.max(PlumPartRevision.revision_number)).where(
            PlumPartRevision.part_id == part_id
        )
    )
    max_num: int = max_num_result.scalar() or 0
    next_number = max_num + 1

    # Compute next draft label
    source_label = source.revision_label if source else _first_revision_label(scheme)
    next_label = _next_draft_label(scheme, source_label)

    # Copy-forward attributes from source; RevisionCreate fields override if provided
    # Phase 6: also copy cost columns forward (D-01: BOM + cost carry forward to new Draft).
    # released_cost_snapshot is NOT copied — it is set only on release (D-14).
    # selected_vendor_link_id is defensively null-checked: if the referenced AVL link
    # was deleted, carrying a stale FK would be invalid; clear it (RESEARCH Open Question 1).
    source_selected_vendor_link_id = None
    source_selected_price_break_index = None
    if source is not None and source.selected_vendor_link_id:
        # Validate that the AVL link still exists before copying forward
        from app.modules.plum.models import PlumAvlLink
        avl_check = await db.execute(
            select(PlumAvlLink.id).where(PlumAvlLink.id == source.selected_vendor_link_id)
        )
        if avl_check.scalar() is not None:
            source_selected_vendor_link_id = source.selected_vendor_link_id
            source_selected_price_break_index = source.selected_price_break_index

    new_revision = PlumPartRevision(
        id=str(uuid.uuid4()),
        part_id=part_id,
        revision_number=next_number,
        revision_label=next_label,
        status="draft",
        description=data.description if data.description is not None else (source.description if source else ""),
        category=data.category if data.category is not None else (source.category if source else None),
        unit_of_measure=data.unit_of_measure if data.unit_of_measure is not None else (source.unit_of_measure if source else None),
        notes=data.notes if data.notes is not None else (source.notes if source else None),
        reason_for_revision=data.reason_for_revision,
        # Phase 6 cost copy-forward (D-01); released_cost_snapshot excluded (D-14)
        material_cost=source.material_cost if source else None,
        sale_price=source.sale_price if source else None,
        selected_vendor_link_id=source_selected_vendor_link_id,
        selected_price_break_index=source_selected_price_break_index,
    )
    db.add(new_revision)
    # Flush so new_revision.id is available for BOM copy-forward below
    await db.flush()

    # Phase 6: copy BOM lines forward from source revision (D-01)
    if source is not None:
        await _copy_bom_forward(db, source.id, new_revision.id)

    await db.commit()
    await db.refresh(new_revision)

    await write_audit(
        db,
        actor_id=actor_id,
        action="revision.created",
        target_type="revision",
        target_id=str(new_revision.id),
        detail=f"Revision {new_revision.revision_label} created for part {part_id}",
    )

    return new_revision


async def advance_revision_status(
    db: AsyncSession,
    part_id: str,
    revision_id: str,
    target_status: str,
    actor_id: str,
) -> "PlumPartRevision":
    """
    Advance a revision through the FSM (D-07/D-08).

    Validates:
      - Part exists (404 if not).
      - Revision belongs to the part (404 if not).
      - Transition is valid per VALID_TRANSITIONS (422 if not).

    On target_status == "released" (D-08 supersede):
      1. Load prior Released revision (if any).
      2. Set it to "obsolete" + obsoleted_at.
      3. await db.flush() between the two updates (Pitfall 3 — partial unique
         index uq_plum_part_one_released would see two "released" rows without flush).
      4. Write revision.obsoleted audit event for the prior revision.
      5. Set this revision to "released" + released_at.

    Commits the transaction; writes the target-specific audit event.
    Returns the updated PlumPartRevision ORM instance.

    Audit actions per target:
      - "in_review" → "revision.submitted"
      - "released"  → "revision.released"
      - "draft"     → "revision.rejected"  (in_review → draft rejection)
      - "obsolete"  → not exposed via API directly (supersede-only path)
    """
    from app.modules.auth.service import write_audit
    from app.modules.plum.models import PlumPartRevision

    # Verify part exists
    await get_part(db, part_id)

    # Load revision + verify ownership
    revision = await get_revision(db, revision_id)
    if revision.part_id != part_id:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Revision {revision_id} does not belong to part {part_id}",
        )

    # Validate FSM transition
    allowed = VALID_TRANSITIONS.get(revision.status, [])
    if target_status not in allowed:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=(
                f"Cannot transition revision from '{revision.status}' to '{target_status}'. "
                f"Allowed transitions: {allowed}"
            ),
        )

    # Supersede on release (D-08)
    if target_status == "released":
        prior_released = await get_released_revision(db, part_id)
        if prior_released and prior_released.id != revision_id:
            prior_released.status = "obsolete"
            prior_released.obsoleted_at = datetime.now(timezone.utc)
            # Flush before the second update to avoid violating the partial unique
            # index uq_plum_part_one_released (Pitfall 3 from RESEARCH.md)
            await db.flush()
            await write_audit(
                db,
                actor_id=actor_id,
                action="revision.obsoleted",
                target_type="revision",
                target_id=str(prior_released.id),
                detail=f"Superseded by revision {revision.revision_label}",
            )

        # Phase 6: D-14 — snapshot the effective cost BEFORE status flip
        # The snapshot must be written before revision.status changes to "released"
        # so compute_effective_cost can still use a consistent state.
        bom_rollup = await _compute_bom_rollup(db, revision.id)
        effective_cost, _ = await compute_effective_cost(db, revision, bom_rollup)
        revision.released_cost_snapshot = effective_cost

        # On release, update the label for SemVer scheme (major bump)
        scheme = await _get_revision_scheme(db)
        new_label = _release_label(scheme, revision.revision_label)
        revision.revision_label = new_label
        revision.released_at = datetime.now(timezone.utc)

    # Apply the transition
    revision.status = target_status

    await db.commit()
    await db.refresh(revision)

    # Write the target-specific audit event
    audit_action_map = {
        "in_review": "revision.submitted",
        "released":  "revision.released",
        "draft":     "revision.rejected",
        "obsolete":  "revision.obsoleted",
    }
    audit_action = audit_action_map.get(target_status, f"revision.{target_status}")
    await write_audit(
        db,
        actor_id=actor_id,
        action=audit_action,
        target_type="revision",
        target_id=str(revision_id),
        detail=f"Revision {revision.revision_label} advanced to {target_status}",
    )

    return revision


# ===========================================================================
# Phase 6: BOM CRUD — add / update / remove BOM lines (PLUM-04/D-01/D-05)
# ===========================================================================


async def _would_create_cycle(
    db: AsyncSession,
    parent_part_id: str,
    candidate_child_id: str,
) -> bool:
    """
    Return True if adding candidate_child_id as a child of parent_part_id
    would create a cycle in the BOM graph (D-05).

    A cycle occurs when parent_part_id is already a descendant (transitively)
    of candidate_child_id. Uses iterative BFS downward from candidate_child_id
    through all BOM lines. (RESEARCH Pattern 9 / plm_v54 checkCircularBom.)

    Trivial case: a part cannot be its own child.
    """
    from app.modules.plum.models import PlumBomItem, PlumPartRevision

    # Trivial self-reference
    if parent_part_id == candidate_child_id:
        return True

    # BFS: traverse descendants of candidate_child_id
    visited: set[str] = set()
    queue: list[str] = [candidate_child_id]

    while queue:
        current = queue.pop(0)
        if current in visited:
            continue
        visited.add(current)

        # Find BOM lines for the latest revision of `current`
        # (we check any revision to be conservative — if parent appears in any
        #  child's known BOM, the graph would cycle)
        latest_rev_sq = (
            select(func.max(PlumPartRevision.revision_number))
            .where(PlumPartRevision.part_id == current)
            .scalar_subquery()
        )
        rev_result = await db.execute(
            select(PlumPartRevision.id).where(
                PlumPartRevision.part_id == current,
                PlumPartRevision.revision_number == latest_rev_sq,
            )
        )
        rev_id = rev_result.scalar()

        if rev_id is None:
            continue

        child_result = await db.execute(
            select(PlumBomItem.child_part_id).where(
                PlumBomItem.parent_revision_id == rev_id
            )
        )
        for (child_id,) in child_result.all():
            if child_id == parent_part_id:
                return True  # cycle detected
            if child_id not in visited:
                queue.append(child_id)

    return False


async def _resolve_child_revision(
    db: AsyncSession,
    child_part_id: str,
) -> tuple:
    """
    Resolve a child part to its effective revision per D-02/D-03.

    Returns (revision, is_unreleased) where:
      is_unreleased=False: child has a Released revision (authoritative, D-02)
      is_unreleased=True:  child has no Released revision; using latest Draft (D-03)
    """
    from app.modules.plum.models import PlumPartRevision

    # Try latest Released revision first (D-02)
    released = await get_released_revision(db, child_part_id)
    if released:
        return (released, False)

    # Fallback: latest overall revision (D-03 provisional)
    latest_result = await db.execute(
        select(PlumPartRevision)
        .where(PlumPartRevision.part_id == child_part_id)
        .order_by(PlumPartRevision.revision_number.desc())
    )
    latest = latest_result.scalars().first()
    return (latest, True)


async def _copy_bom_forward(
    db: AsyncSession,
    source_revision_id: str,
    new_revision_id: str,
) -> int:
    """
    Copy all BOM lines from source_revision to new_revision.

    Used in create_revision (D-01: BOM copies forward to new Draft).
    Uses db.flush() (not commit) — called inside create_revision's transaction.
    Returns count of lines copied.
    """
    from app.modules.plum.models import PlumBomItem

    existing = await db.execute(
        select(PlumBomItem)
        .where(PlumBomItem.parent_revision_id == source_revision_id)
        .order_by(PlumBomItem.sort_order)
    )
    lines = list(existing.scalars().all())

    for line in lines:
        db.add(
            PlumBomItem(
                id=str(uuid.uuid4()),
                parent_revision_id=new_revision_id,
                child_part_id=line.child_part_id,
                qty=line.qty,
                ref_des=line.ref_des,
                sort_order=line.sort_order,
            )
        )

    # flush (not commit) — stay inside the parent create_revision transaction
    if lines:
        await db.flush()
    return len(lines)


async def add_bom_line(
    db: AsyncSession,
    part_id: str,
    data: "BomItemCreate",
    revision_id: str,
    actor_id: str,
) -> "PlumBomItem":
    """
    Add a child part to a Draft revision's BOM (PLUM-04/D-01/D-04).

    Raises:
      HTTP 404 if revision not found or doesn't belong to part_id.
      HTTP 422 if the revision is not in "draft" status (D-01 immutability, T-06-05).
      HTTP 422 if adding child_part_id would create a cycle (D-05, T-06-06).
      HTTP 409 via DB UniqueConstraint if the same child already exists (T-06-03).

    Writes a bom.line_added audit event.
    Returns the new PlumBomItem ORM instance.
    """
    from app.modules.auth.service import write_audit
    from app.modules.plum.models import PlumBomItem, PlumPart

    revision = await get_revision(db, revision_id)
    if revision.part_id != part_id:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Revision {revision_id} does not belong to part {part_id}",
        )

    if revision.status != "draft":
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="BOM lines can only be edited on Draft revisions.",
        )

    # Verify child part exists
    child_result = await db.execute(
        select(PlumPart).where(PlumPart.id == data.child_part_id)
    )
    child_part = child_result.scalars().first()
    if child_part is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Child part {data.child_part_id} not found",
        )

    # D-05: cycle detection (T-06-06)
    if await _would_create_cycle(db, revision.part_id, data.child_part_id):
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=(
                f"Adding {child_part.part_number} here would create a circular BOM. "
                "Choose a different part."
            ),
        )

    # Determine next sort order
    max_sort_result = await db.execute(
        select(func.max(PlumBomItem.sort_order)).where(
            PlumBomItem.parent_revision_id == revision_id
        )
    )
    next_sort = (max_sort_result.scalar() or 0) + 1
    if data.sort_order is not None:
        next_sort = data.sort_order

    item = PlumBomItem(
        id=str(uuid.uuid4()),
        parent_revision_id=revision_id,
        child_part_id=data.child_part_id,
        qty=data.qty,
        ref_des=data.ref_des,
        sort_order=next_sort,
    )
    db.add(item)
    await db.commit()
    await db.refresh(item)

    await write_audit(
        db,
        actor_id=actor_id,
        action="bom.line_added",
        target_type="bom_item",
        target_id=str(item.id),
        detail=f"BOM line added to revision {revision_id}: child {data.child_part_id} qty={data.qty}",
    )

    return item


async def update_bom_line(
    db: AsyncSession,
    part_id: str,
    line_id: str,
    data: "BomItemUpdate",
    actor_id: str,
) -> "PlumBomItem":
    """
    Update a BOM line's qty / ref_des / sort_order (PLUM-04/D-01).

    Raises HTTP 422 if the parent revision is not Draft (D-01 immutability).
    Raises HTTP 404 if line not found or not owned by part_id.
    Writes a bom.line_updated audit event.
    """
    from app.modules.auth.service import write_audit
    from app.modules.plum.models import PlumBomItem

    line_result = await db.execute(
        select(PlumBomItem).where(PlumBomItem.id == line_id)
    )
    line = line_result.scalars().first()

    if line is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"BOM line {line_id} not found",
        )

    # Load the parent revision to check ownership + immutability
    revision = await get_revision(db, line.parent_revision_id)
    if revision.part_id != part_id:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"BOM line {line_id} does not belong to part {part_id}",
        )

    if revision.status != "draft":
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="BOM lines can only be edited on Draft revisions.",
        )

    update_data = data.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(line, field, value)

    await db.commit()
    await db.refresh(line)

    await write_audit(
        db,
        actor_id=actor_id,
        action="bom.line_updated",
        target_type="bom_item",
        target_id=str(line_id),
        detail=f"BOM line updated on revision {line.parent_revision_id}",
    )

    return line


async def remove_bom_line(
    db: AsyncSession,
    part_id: str,
    line_id: str,
    actor_id: str,
) -> None:
    """
    Remove a BOM line from a Draft revision (PLUM-04/D-01).

    Raises HTTP 422 if the parent revision is not Draft.
    Raises HTTP 404 if line not found or not owned by part_id.
    Writes a bom.line_removed audit event.
    """
    from app.modules.auth.service import write_audit
    from app.modules.plum.models import PlumBomItem

    line_result = await db.execute(
        select(PlumBomItem).where(PlumBomItem.id == line_id)
    )
    line = line_result.scalars().first()

    if line is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"BOM line {line_id} not found",
        )

    revision = await get_revision(db, line.parent_revision_id)
    if revision.part_id != part_id:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"BOM line {line_id} does not belong to part {part_id}",
        )

    if revision.status != "draft":
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="BOM lines can only be edited on Draft revisions.",
        )

    await db.delete(line)
    await db.commit()

    await write_audit(
        db,
        actor_id=actor_id,
        action="bom.line_removed",
        target_type="bom_item",
        target_id=str(line_id),
        detail=f"BOM line removed from revision {revision.id}",
    )


# ===========================================================================
# Phase 6: BOM traversal — tree / flat / where-used (PLUM-04/05/06)
# ===========================================================================


async def _build_bom_tree_recursive(
    db: AsyncSession,
    revision_id: str,
    depth: int,
    visited_revisions: set[str],
) -> list[dict]:
    """
    Recursively build BOM tree nodes for a given revision.

    visited_revisions is copied per branch (visited.copy()) to allow a part to
    appear in multiple branches without false cycle detection (the visited set
    tracks the current path, not all parts seen globally).

    Each child is resolved to its latest Released revision (D-02), or latest
    Draft (D-03) if no Released exists.

    Returns a list of dicts shaped for BomTreeNode serialization.
    """
    from app.modules.plum.models import PlumBomItem, PlumPart

    if revision_id in visited_revisions or depth > 20:
        return []

    # Mark this revision as visited on THIS branch
    branch_visited = visited_revisions.copy()
    branch_visited.add(revision_id)

    # Fetch BOM lines for this revision, ordered by sort_order
    lines_result = await db.execute(
        select(PlumBomItem)
        .where(PlumBomItem.parent_revision_id == revision_id)
        .order_by(PlumBomItem.sort_order)
    )
    lines = list(lines_result.scalars().all())

    nodes = []
    for line in lines:
        child_revision, is_unreleased = await _resolve_child_revision(db, line.child_part_id)

        # Load child part info for part_number / uom
        part_result = await db.execute(
            select(PlumPart).where(PlumPart.id == line.child_part_id)
        )
        child_part = part_result.scalars().first()
        if child_part is None:
            continue

        unit_of_measure = child_revision.unit_of_measure if child_revision else None

        # Recurse into child's BOM
        child_rev_id = child_revision.id if child_revision else None
        children = []
        if child_rev_id:
            children = await _build_bom_tree_recursive(
                db, child_rev_id, depth + 1, branch_visited
            )

        nodes.append(
            {
                "bom_item_id": line.id,
                "part_id": line.child_part_id,
                "part_number": child_part.part_number,
                "unit_of_measure": unit_of_measure,
                "qty": line.qty,
                "ref_des": line.ref_des,
                "sort_order": line.sort_order,
                "depth": depth,
                "is_unreleased": is_unreleased,
                "effective_cost": None,  # populated by caller if needed
                "effective_cost_source": None,
                "children": children,
            }
        )

    return nodes


async def load_bom_tree(
    db: AsyncSession,
    part_id: str,
    revision_id: str,
) -> list[dict]:
    """
    Load the BOM as a recursive tree (PLUM-04/D-02/D-03).

    Returns a list of top-level BomTreeNode dicts. Each node has `children`
    (recursively nested). Child parts are resolved to their latest Released
    revision (D-02), falling back to latest Draft with is_unreleased=True (D-03).
    Max depth guard: 20 levels (prevents runaway recursion on pathological inputs).

    Raises HTTP 404 if the revision doesn't belong to part_id.
    """
    revision = await get_revision(db, revision_id)
    if revision.part_id != part_id:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Revision {revision_id} does not belong to part {part_id}",
        )

    return await _build_bom_tree_recursive(db, revision_id, depth=0, visited_revisions=set())


async def _compute_bom_rollup(
    db: AsyncSession,
    revision_id: str,
    visited_revisions: set[str] | None = None,
) -> Decimal | None:
    """
    Recursively compute the BOM roll-up cost for a revision (D-08).

    Returns Σ(child effective_cost × line.qty) across all direct children,
    or None if no children have costs.
    """
    from app.modules.plum.models import PlumBomItem

    if visited_revisions is None:
        visited_revisions = set()

    if revision_id in visited_revisions:
        return None  # cycle guard (shouldn't happen — D-05 prevents cycles)

    visited_revisions = visited_revisions.copy()
    visited_revisions.add(revision_id)

    lines_result = await db.execute(
        select(PlumBomItem).where(PlumBomItem.parent_revision_id == revision_id)
    )
    lines = list(lines_result.scalars().all())

    if not lines:
        return None

    total: Decimal | None = None
    for line in lines:
        child_revision, _ = await _resolve_child_revision(db, line.child_part_id)
        if child_revision is None:
            continue

        # Recursively compute child's BOM roll-up (for step 3 fallback)
        child_rollup = await _compute_bom_rollup(db, child_revision.id, visited_revisions)
        child_cost, _ = await compute_effective_cost(db, child_revision, child_rollup)

        if child_cost is not None:
            contribution = Decimal(str(child_cost)) * Decimal(str(line.qty))
            total = (total or Decimal("0")) + contribution

    return total


async def load_flat_bom(
    db: AsyncSession,
    part_id: str,
    revision_id: str,
) -> list[dict]:
    """
    Load a flat BOM with total quantity rolled up across all paths (PLUM-05/D-04).

    Uses a dict accumulator keyed by child_part_id. When the same part appears
    in multiple sub-assemblies, its total_qty is the sum of (path qty products)
    across all paths (Pitfall 8: cumulative_qty multiplication down each path).

    Returns a list of FlatBomRow dicts sorted by part_number.
    Raises HTTP 404 if revision doesn't belong to part_id.
    """
    from app.modules.plum.models import PlumBomItem, PlumPart

    revision = await get_revision(db, revision_id)
    if revision.part_id != part_id:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Revision {revision_id} does not belong to part {part_id}",
        )

    # Accumulator: part_id → dict with total_qty + metadata
    flat: dict[str, dict] = {}

    async def _accumulate(rev_id: str, path_qty: Decimal, visited: set[str]) -> None:
        if rev_id in visited:
            return
        branch_visited = visited.copy()
        branch_visited.add(rev_id)

        lines_result = await db.execute(
            select(PlumBomItem)
            .where(PlumBomItem.parent_revision_id == rev_id)
        )
        lines = list(lines_result.scalars().all())

        for line in lines:
            child_part_id = line.child_part_id
            # Cumulative qty at this level = path_qty × line.qty
            cumulative_qty = path_qty * Decimal(str(line.qty))

            # Load part info (cached implicitly via dict keying)
            if child_part_id not in flat:
                part_result = await db.execute(
                    select(PlumPart).where(PlumPart.id == child_part_id)
                )
                child_part = part_result.scalars().first()
                if child_part is None:
                    continue

                child_rev, is_unreleased = await _resolve_child_revision(db, child_part_id)
                flat[child_part_id] = {
                    "part_id": child_part_id,
                    "part_number": child_part.part_number,
                    "unit_of_measure": child_rev.unit_of_measure if child_rev else None,
                    "total_qty": Decimal("0"),
                    "effective_cost": None,
                    "extended_cost": None,
                    "is_unreleased": is_unreleased,
                    "_child_rev": child_rev,
                }

            flat[child_part_id]["total_qty"] += cumulative_qty

            # Recurse into child's BOM
            child_rev = flat[child_part_id]["_child_rev"]
            if child_rev:
                await _accumulate(child_rev.id, cumulative_qty, branch_visited)

    await _accumulate(revision_id, Decimal("1"), set())

    # Build final list (remove internal _child_rev key)
    rows = []
    for row_dict in flat.values():
        child_rev = row_dict.pop("_child_rev", None)
        if child_rev:
            child_rollup = await _compute_bom_rollup(db, child_rev.id)
            child_cost, _ = await compute_effective_cost(db, child_rev, child_rollup)
            row_dict["effective_cost"] = child_cost
            if child_cost is not None:
                row_dict["extended_cost"] = Decimal(str(child_cost)) * row_dict["total_qty"]
        rows.append(row_dict)

    # Sort by part_number for stable output
    rows.sort(key=lambda r: r["part_number"])
    return rows


async def get_where_used(
    db: AsyncSession,
    part_id: str,
    max_depth: int = 20,
) -> list[dict]:
    """
    Return direct and indirect parent assemblies that reference this part (PLUM-06).

    Traverses the BOM graph in reverse (upward) from part_id. Direct parents
    are depth=1; indirect (grandparents etc.) are depth>1 with indirect=True
    and via_part_number naming the intermediate part they are reached through.

    Returns a list of WhereUsedRow dicts. Deduplicates by parent_part_id
    (a part may appear at multiple depths — only the shallowest is kept).
    """
    from app.modules.plum.models import PlumBomItem, PlumPart, PlumPartRevision

    # BFS upward: find revisions whose BOM includes part_id as a child
    found: dict[str, dict] = {}  # parent_part_id → WhereUsedRow dict
    # Queue: (part_id_to_search_up_from, depth, that part's number). The part
    # number rides along so an indirect parent can name the intermediate part it
    # is reached through — it is exactly the part we are searching up from.
    queue: list[tuple[str, int, str | None]] = [(part_id, 0, None)]
    searched_parts: set[str] = {part_id}

    while queue:
        current_part_id, current_depth, current_part_number = queue.pop(0)
        next_depth = current_depth + 1

        if next_depth > max_depth:
            continue

        is_indirect = next_depth > 1

        # Find all BOM lines that have current_part_id as child
        lines_result = await db.execute(
            select(PlumBomItem.parent_revision_id)
            .where(PlumBomItem.child_part_id == current_part_id)
        )
        rev_ids = [row[0] for row in lines_result.all()]

        for rev_id in rev_ids:
            # Load the revision to get its part_id
            rev_result = await db.execute(
                select(PlumPartRevision).where(PlumPartRevision.id == rev_id)
            )
            rev = rev_result.scalars().first()
            if rev is None:
                continue

            parent_part_id = rev.part_id

            # Load parent part info
            part_result = await db.execute(
                select(PlumPart).where(PlumPart.id == parent_part_id)
            )
            parent_part = part_result.scalars().first()
            if parent_part is None:
                continue

            # Record if not already found (keep shallowest depth)
            if parent_part_id not in found:
                found[parent_part_id] = {
                    "parent_part_id": parent_part_id,
                    "parent_part_number": parent_part.part_number,
                    "parent_revision_id": rev_id,
                    "parent_revision_label": rev.revision_label,
                    "parent_revision_status": rev.status,
                    "direct": not is_indirect,
                    "indirect": is_indirect,
                    "via_part_number": current_part_number if is_indirect else None,
                }

            # Continue searching upward from this parent if not already searched
            if parent_part_id not in searched_parts:
                searched_parts.add(parent_part_id)
                queue.append((parent_part_id, next_depth, parent_part.part_number))

    return list(found.values())


# ===========================================================================
# Phase 6: AVL CRUD — Approved Vendor List (PLUM-07/D-11/D-12/D-13)
# ===========================================================================


async def list_avl_links(
    db: AsyncSession,
    part_id: str,
) -> list[dict]:
    """
    Return all AVL links for a part, including their price breaks (PLUM-07/D-11).

    Active=False links are included (archived vendors still surface per Pitfall 4).
    Price breaks are sorted by qty_threshold ascending.
    Returns list of AvlLinkRead dicts.
    """
    from app.modules.plum.models import PlumAvlLink, PlumAvlPriceBreak

    links_result = await db.execute(
        select(PlumAvlLink)
        .where(PlumAvlLink.part_id == part_id)
        .order_by(PlumAvlLink.created_at)
    )
    links = list(links_result.scalars().all())

    result = []
    for link in links:
        pbs_result = await db.execute(
            select(PlumAvlPriceBreak)
            .where(PlumAvlPriceBreak.avl_link_id == link.id)
            .order_by(PlumAvlPriceBreak.qty_threshold)
        )
        price_breaks = list(pbs_result.scalars().all())

        result.append(
            {
                "id": link.id,
                "part_id": link.part_id,
                "vendor_id": link.vendor_id,
                "vendor_part_number": link.vendor_part_number,
                "preferred": link.preferred,
                "notes": link.notes,
                "active": link.active,
                "price_breaks": [
                    {
                        "id": pb.id,
                        "avl_link_id": pb.avl_link_id,
                        "qty_threshold": pb.qty_threshold,
                        "unit_cost": pb.unit_cost,
                        "lead_days": pb.lead_days,
                        "sort_order": pb.sort_order,
                    }
                    for pb in price_breaks
                ],
            }
        )

    return result


async def add_avl_link(
    db: AsyncSession,
    part_id: str,
    data: "AvlLinkCreate",
    actor_id: str,
) -> dict:
    """
    Link a part to a SYERP vendor in the AVL (PLUM-07/D-11/D-13).

    Validates that the vendor exists and is_vendor=True (T-06-07).
    Returns AvlLinkRead dict (with empty price_breaks list — add via separate endpoint).
    Writes a avl.link_added audit event.

    Raises HTTP 422 if vendor not found or is_vendor=False.
    """
    from app.modules.auth.service import write_audit
    from app.modules.plum.models import PlumAvlLink
    from app.modules.syerp.models import Partner as SyerpPartner

    # Validate vendor is_vendor=True (T-06-07)
    vendor_result = await db.execute(
        select(SyerpPartner).where(
            SyerpPartner.id == data.vendor_id,
            SyerpPartner.is_vendor == True,  # noqa: E712
        )
    )
    vendor = vendor_result.scalars().first()
    if vendor is None:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Vendor not found or not a vendor.",
        )

    link = PlumAvlLink(
        id=str(uuid.uuid4()),
        part_id=part_id,
        vendor_id=data.vendor_id,
        vendor_part_number=data.vendor_part_number,
        preferred=data.preferred,
        notes=data.notes,
        active=True,
    )
    db.add(link)
    await db.commit()
    await db.refresh(link)

    await write_audit(
        db,
        actor_id=actor_id,
        action="avl.link_added",
        target_type="avl_link",
        target_id=str(link.id),
        detail=f"AVL link added: part {part_id} → vendor {data.vendor_id}",
    )

    return {
        "id": link.id,
        "part_id": link.part_id,
        "vendor_id": link.vendor_id,
        "vendor_part_number": link.vendor_part_number,
        "preferred": link.preferred,
        "notes": link.notes,
        "active": link.active,
        "price_breaks": [],
    }


async def update_avl_link(
    db: AsyncSession,
    part_id: str,
    link_id: str,
    data: "AvlLinkUpdate",
    actor_id: str,
) -> dict:
    """
    Update an AVL link's metadata (PLUM-07/D-11).

    Does NOT update vendor_id or price_breaks (use separate endpoints).
    Writes a avl.link_updated audit event.
    Returns updated AvlLinkRead dict.
    """
    from app.modules.auth.service import write_audit
    from app.modules.plum.models import PlumAvlLink

    link_result = await db.execute(
        select(PlumAvlLink).where(
            PlumAvlLink.id == link_id,
            PlumAvlLink.part_id == part_id,
        )
    )
    link = link_result.scalars().first()
    if link is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"AVL link {link_id} not found for part {part_id}",
        )

    update_data = data.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(link, field, value)

    await db.commit()
    await db.refresh(link)

    await write_audit(
        db,
        actor_id=actor_id,
        action="avl.link_updated",
        target_type="avl_link",
        target_id=str(link_id),
        detail=f"AVL link updated: part {part_id} link {link_id}",
    )

    # Return current link dict with price breaks
    links = await list_avl_links(db, part_id)
    for lnk in links:
        if lnk["id"] == link_id:
            return lnk
    # Fallback (should not happen)
    return {
        "id": link.id,
        "part_id": link.part_id,
        "vendor_id": link.vendor_id,
        "vendor_part_number": link.vendor_part_number,
        "preferred": link.preferred,
        "notes": link.notes,
        "active": link.active,
        "price_breaks": [],
    }


async def remove_avl_link(
    db: AsyncSession,
    part_id: str,
    link_id: str,
    actor_id: str,
) -> None:
    """
    Remove an AVL link (soft-delete: set active=False) (PLUM-07/D-11).

    Writes a avl.link_removed audit event.
    """
    from app.modules.auth.service import write_audit
    from app.modules.plum.models import PlumAvlLink

    link_result = await db.execute(
        select(PlumAvlLink).where(
            PlumAvlLink.id == link_id,
            PlumAvlLink.part_id == part_id,
        )
    )
    link = link_result.scalars().first()
    if link is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"AVL link {link_id} not found for part {part_id}",
        )

    link.active = False
    await db.commit()

    await write_audit(
        db,
        actor_id=actor_id,
        action="avl.link_removed",
        target_type="avl_link",
        target_id=str(link_id),
        detail=f"AVL link removed (soft-deleted): part {part_id} link {link_id}",
    )


async def add_price_break(
    db: AsyncSession,
    part_id: str,
    link_id: str,
    data: "PriceBreakCreate",
    actor_id: str,
) -> dict:
    """
    Add a price break to an AVL link (PLUM-07/D-11).

    Price breaks are always sorted by qty_threshold ascending on retrieval.
    Returns the new PlumAvlPriceBreak as a dict.
    """
    from app.modules.auth.service import write_audit
    from app.modules.plum.models import PlumAvlLink, PlumAvlPriceBreak

    # Verify link exists and belongs to part
    link_result = await db.execute(
        select(PlumAvlLink).where(
            PlumAvlLink.id == link_id,
            PlumAvlLink.part_id == part_id,
        )
    )
    link = link_result.scalars().first()
    if link is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"AVL link {link_id} not found for part {part_id}",
        )

    # Determine sort_order (append at end by default)
    max_sort_result = await db.execute(
        select(func.max(PlumAvlPriceBreak.sort_order)).where(
            PlumAvlPriceBreak.avl_link_id == link_id
        )
    )
    next_sort = (max_sort_result.scalar() or 0) + 1
    if data.sort_order is not None:
        next_sort = data.sort_order

    pb = PlumAvlPriceBreak(
        id=str(uuid.uuid4()),
        avl_link_id=link_id,
        qty_threshold=data.qty_threshold,
        unit_cost=data.unit_cost,
        lead_days=data.lead_days,
        sort_order=next_sort,
    )
    db.add(pb)
    await db.commit()
    await db.refresh(pb)

    await write_audit(
        db,
        actor_id=actor_id,
        action="avl.price_break_added",
        target_type="avl_price_break",
        target_id=str(pb.id),
        detail=f"Price break added to AVL link {link_id}: qty≥{data.qty_threshold} cost={data.unit_cost}",
    )

    return {
        "id": pb.id,
        "avl_link_id": pb.avl_link_id,
        "qty_threshold": pb.qty_threshold,
        "unit_cost": pb.unit_cost,
        "lead_days": pb.lead_days,
        "sort_order": pb.sort_order,
    }


# ===========================================================================
# Phase 6: Effective-cost chain + margin (PLUM-08/09/D-07/D-09)
# ===========================================================================


async def compute_effective_cost(
    db: AsyncSession,
    revision: "PlumPartRevision",
    bom_rollup: Decimal | None = None,
) -> tuple[Decimal | None, str]:
    """
    Resolve the effective cost for a revision per D-07 priority chain.

    Returns (cost_value, source_label) where source_label is one of:
      "vendor price" | "manual" | "roll-up" | "uncosted"

    Priority:
      1. Selected vendor + price-break unit_cost (if selected_vendor_link_id set)
         — bounds-checked against actual list length (Pitfall 7 in RESEARCH.md)
      2. Manual material_cost (if set on revision)
      3. BOM roll-up (passed as bom_rollup argument by caller)
      4. None / "uncosted"
    """
    from app.modules.plum.models import PlumAvlPriceBreak

    # Step 1: Selected vendor price-break
    if (
        revision.selected_vendor_link_id is not None
        and revision.selected_price_break_index is not None
    ):
        pb_result = await db.execute(
            select(PlumAvlPriceBreak)
            .where(PlumAvlPriceBreak.avl_link_id == revision.selected_vendor_link_id)
            .order_by(PlumAvlPriceBreak.qty_threshold)
        )
        price_breaks = list(pb_result.scalars().all())
        idx = revision.selected_price_break_index
        # Bounds-check: if index is out of range, fall through (Pitfall 7)
        if 0 <= idx < len(price_breaks):
            return (Decimal(str(price_breaks[idx].unit_cost)), "vendor price")

    # Step 2: Manual material cost
    if revision.material_cost is not None:
        return (Decimal(str(revision.material_cost)), "manual")

    # Step 3: BOM roll-up (caller pre-computes and passes in)
    if bom_rollup is not None:
        return (Decimal(str(bom_rollup)), "roll-up")

    # Step 4: Uncosted
    return (None, "uncosted")


def compute_margin(
    effective_cost: Decimal | None,
    sale_price: Decimal | None,
) -> tuple[Decimal | None, Decimal | None]:
    """
    Compute margin and margin % for display (PLUM-09/D-09).

    margin = sale_price − effective_cost
    margin_pct = margin / effective_cost × 100

    Returns (margin, margin_pct). Both are None if either input is None.
    Negative margin is allowed (returned, not rejected — D-09).
    """
    if effective_cost is None or sale_price is None:
        return (None, None)

    margin = Decimal(str(sale_price)) - Decimal(str(effective_cost))
    if effective_cost == 0:
        margin_pct = None  # division by zero — cost is zero
    else:
        margin_pct = (margin / Decimal(str(effective_cost))) * Decimal("100")

    return (margin, margin_pct)


async def update_cost(
    db: AsyncSession,
    part_id: str,
    revision_id: str,
    data: "CostUpdate",
    actor_id: str,
) -> "PlumPartRevision":
    """
    Update cost fields on a Draft revision (PLUM-08/D-06/D-07).

    Raises HTTP 422 if the revision is not in "draft" status (D-01 immutability).
    Updates material_cost, sale_price, selected_vendor_link_id,
    selected_price_break_index on the revision.
    Writes a part.cost_updated audit event.
    Returns the updated PlumPartRevision ORM instance.
    """
    from app.modules.auth.service import write_audit

    revision = await get_revision(db, revision_id)
    if revision.part_id != part_id:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Revision {revision_id} does not belong to part {part_id}",
        )

    if revision.status != "draft":
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="BOM lines can only be edited on Draft revisions.",
        )

    update_data = data.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(revision, field, value)

    await db.commit()
    await db.refresh(revision)

    await write_audit(
        db,
        actor_id=actor_id,
        action="part.cost_updated",
        target_type="revision",
        target_id=str(revision_id),
        detail=f"Cost fields updated on revision {revision_id}",
    )

    return revision


async def get_cost_read(
    db: AsyncSession,
    part_id: str,
    revision_id: str,
) -> dict:
    """
    Return the full CostRead dict for a revision (PLUM-08/09).

    Computes:
      - bom_rollup_cost: live BOM roll-up (always computed, even for Released revisions)
      - effective_cost + effective_cost_source via D-07 chain
      - margin + margin_pct via compute_margin
      - released_cost_snapshot: the frozen cost snapshot (if revision is Released)
    """
    revision = await get_revision(db, revision_id)
    if revision.part_id != part_id:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Revision {revision_id} does not belong to part {part_id}",
        )

    # Always compute live BOM roll-up (D-14: shown alongside frozen snapshot)
    bom_rollup = await _compute_bom_rollup(db, revision_id)
    effective_cost, effective_cost_source = await compute_effective_cost(
        db, revision, bom_rollup
    )
    margin, margin_pct = compute_margin(effective_cost, revision.sale_price)

    return {
        "material_cost": revision.material_cost,
        "sale_price": revision.sale_price,
        "released_cost_snapshot": revision.released_cost_snapshot,
        "selected_vendor_link_id": revision.selected_vendor_link_id,
        "selected_price_break_index": revision.selected_price_break_index,
        "effective_cost": effective_cost,
        "effective_cost_source": effective_cost_source,
        "bom_rollup_cost": bom_rollup,
        "margin": margin,
        "margin_pct": margin_pct,
    }


# ---------------------------------------------------------------------------
# Import / Export (PLUM-10, D-15/D-16/D-17/D-18)
# ---------------------------------------------------------------------------
# JSON schema (D-16 lossless round-trip):
#   { schema_version: 1, exported_at: ISO-8601, parts: [
#     { part_number, active, tags, revisions: [
#       { revision_label, revision_number, status, description,
#         material_cost, sale_price, released_cost_snapshot,
#         selected_vendor_link_id, selected_price_break_index,
#         bom: [{ child_part_number, quantity, reference_designators }]
#       }],
#       avl: [{ vendor_code, vendor_part_number, preferred, notes,
#               price_breaks: [{ qty_threshold, unit_cost, lead_days }] }]
#     }
#   ]}
# Costs serialized as strings from Decimal (lossless, no float).
# ---------------------------------------------------------------------------


async def build_json_export(db: AsyncSession) -> dict:
    """
    Build the full lossless JSON export dataset (PLUM-10, D-16).

    Serializes every PlumPart with all revisions, BOM lines, costs (raw stored
    columns only — NOT live effective_cost), AVL links, and price breaks.
    Decimal values are cast to str for lossless JSON round-trip.
    Returns a dict with schema_version=1 and an exported_at timestamp.
    """
    from app.modules.plum.models import (
        PlumAvlLink,
        PlumAvlPriceBreak,
        PlumBomItem,
        PlumPart,
        PlumPartRevision,
    )
    from sqlalchemy import text as sa_text

    # Load all parts
    parts_result = await db.execute(
        select(PlumPart).order_by(PlumPart.part_number)
    )
    all_parts = list(parts_result.scalars().all())

    # Load all tags (part_id -> list of tag names)
    tags_result = await db.execute(
        sa_text(
            """
            SELECT pt.part_id, ct.name
            FROM plum_part_tag pt
            JOIN plum_classification_tag ct ON ct.id = pt.tag_id
            ORDER BY pt.part_id, ct.name
            """
        )
    )
    tags_by_part: dict[str, list[str]] = {}
    for row in tags_result.all():
        pid = str(row[0])
        tags_by_part.setdefault(pid, []).append(str(row[1]))

    # Load all revisions
    revs_result = await db.execute(
        select(PlumPartRevision).order_by(
            PlumPartRevision.part_id, PlumPartRevision.revision_number
        )
    )
    all_revisions = list(revs_result.scalars().all())
    revs_by_part: dict[str, list] = {}
    for rev in all_revisions:
        revs_by_part.setdefault(str(rev.part_id), []).append(rev)

    # Load all BOM items
    bom_result = await db.execute(
        select(PlumBomItem).order_by(
            PlumBomItem.parent_revision_id, PlumBomItem.sort_order
        )
    )
    all_bom_items = list(bom_result.scalars().all())
    bom_by_rev: dict[str, list] = {}
    for bi in all_bom_items:
        bom_by_rev.setdefault(str(bi.parent_revision_id), []).append(bi)

    # Build part_number lookup from part_id
    part_number_by_id: dict[str, str] = {
        str(p.id): str(p.part_number) for p in all_parts
    }

    # Load all AVL links
    avl_result = await db.execute(
        select(PlumAvlLink).order_by(PlumAvlLink.part_id)
    )
    all_avl_links = list(avl_result.scalars().all())
    avl_by_part: dict[str, list] = {}
    for link in all_avl_links:
        avl_by_part.setdefault(str(link.part_id), []).append(link)

    # Load all price breaks
    pb_result = await db.execute(
        select(PlumAvlPriceBreak).order_by(
            PlumAvlPriceBreak.avl_link_id, PlumAvlPriceBreak.qty_threshold
        )
    )
    all_price_breaks = list(pb_result.scalars().all())
    pb_by_link: dict[str, list] = {}
    for pb in all_price_breaks:
        pb_by_link.setdefault(str(pb.avl_link_id), []).append(pb)

    # Resolve vendor_code from syerp_partner
    vendor_ids = {str(link.vendor_id) for link in all_avl_links}
    vendor_code_by_id: dict[str, str] = {}
    if vendor_ids:
        from app.modules.syerp.models import Partner as SyerpPartner

        vend_result = await db.execute(
            select(SyerpPartner.id, SyerpPartner.code).where(
                SyerpPartner.id.in_(list(vendor_ids))
            )
        )
        for vid, vcode in vend_result.all():
            vendor_code_by_id[str(vid)] = str(vcode)

    # Assemble the export payload
    parts_payload: list[dict] = []
    for part in all_parts:
        pid = str(part.id)
        # Revisions
        revisions_payload: list[dict] = []
        for rev in revs_by_part.get(pid, []):
            rid = str(rev.id)
            # BOM lines for this revision
            bom_payload: list[dict] = []
            for bi in bom_by_rev.get(rid, []):
                child_pn = part_number_by_id.get(str(bi.child_part_id), "")
                bom_payload.append(
                    {
                        "child_part_number": child_pn,
                        "quantity": str(bi.qty) if bi.qty is not None else None,
                        "reference_designators": bi.ref_des,
                    }
                )
            revisions_payload.append(
                {
                    "revision_label": rev.revision_label,
                    "revision_number": rev.revision_number,
                    "status": rev.status,
                    "description": rev.description,
                    "material_cost": (
                        str(rev.material_cost) if rev.material_cost is not None else None
                    ),
                    "sale_price": (
                        str(rev.sale_price) if rev.sale_price is not None else None
                    ),
                    "released_cost_snapshot": (
                        str(rev.released_cost_snapshot)
                        if rev.released_cost_snapshot is not None
                        else None
                    ),
                    "selected_vendor_link_id": rev.selected_vendor_link_id,
                    "selected_price_break_index": rev.selected_price_break_index,
                    "bom": bom_payload,
                }
            )
        # AVL links
        avl_payload: list[dict] = []
        for link in avl_by_part.get(pid, []):
            lid = str(link.id)
            pb_payload: list[dict] = []
            for pb in pb_by_link.get(lid, []):
                pb_payload.append(
                    {
                        "qty_threshold": pb.qty_threshold,
                        "unit_cost": (
                            str(pb.unit_cost) if pb.unit_cost is not None else None
                        ),
                        "lead_days": pb.lead_days,
                    }
                )
            avl_payload.append(
                {
                    "vendor_code": vendor_code_by_id.get(str(link.vendor_id), ""),
                    "vendor_part_number": link.vendor_part_number,
                    "preferred": link.preferred,
                    "notes": link.notes,
                    "price_breaks": pb_payload,
                }
            )
        parts_payload.append(
            {
                "part_number": str(part.part_number),
                "active": bool(part.active),
                "tags": tags_by_part.get(pid, []),
                "revisions": revisions_payload,
                "avl": avl_payload,
            }
        )

    return {
        "schema_version": 1,
        "exported_at": datetime.now(timezone.utc).isoformat(),
        "parts": parts_payload,
    }


def generate_excel_export(data: dict) -> bytes:
    """
    Generate a three-sheet .xlsx workbook from the export data dict (PLUM-10, D-16).

    Sheets:
      "Parts" — part_number, description (from latest revision), tags
      "BOMs"  — parent_part_number, parent_revision_label, child_part_number, quantity,
                 reference_designators
      "AVL"   — part_number, vendor_code, vendor_part_number, preferred, notes,
                 pb_qty_threshold, pb_unit_cost, pb_lead_days (one row per price break)

    Uses openpyxl; workbook saved to BytesIO and returned as bytes.
    T-06-12 mitigation: no formula execution — only data written via .append().
    """
    from io import BytesIO
    import openpyxl

    wb = openpyxl.Workbook()

    # --- Sheet 1: Parts ---
    ws_parts = wb.active
    ws_parts.title = "Parts"
    ws_parts.append(
        ["part_number", "description", "category", "unit_of_measure", "tags", "notes"]
    )
    for part in data.get("parts", []):
        # Extract description from the latest revision (highest revision_number)
        revisions = part.get("revisions", [])
        description = ""
        if revisions:
            latest = max(revisions, key=lambda r: r.get("revision_number", 0))
            description = latest.get("description", "") or ""
        tags = ",".join(part.get("tags", []))
        ws_parts.append(
            [
                part.get("part_number", ""),
                description,
                "",  # category not in export JSON; column kept for human-friendly template
                "",  # unit_of_measure — on revision, not part
                tags,
                "",  # notes — not present in current schema
            ]
        )

    # --- Sheet 2: BOMs (sheet named "BOM" per test_import_export Wave-0 stub) ---
    ws_boms = wb.create_sheet("BOM")
    ws_boms.append(
        [
            "parent_part_number",
            "parent_revision_label",
            "child_part_number",
            "quantity",
            "reference_designators",
        ]
    )
    for part in data.get("parts", []):
        parent_pn = part.get("part_number", "")
        for rev in part.get("revisions", []):
            rev_label = rev.get("revision_label", "")
            for bom_line in rev.get("bom", []):
                ws_boms.append(
                    [
                        parent_pn,
                        rev_label,
                        bom_line.get("child_part_number", ""),
                        bom_line.get("quantity", ""),
                        bom_line.get("reference_designators", ""),
                    ]
                )

    # --- Sheet 3: AVL ---
    ws_avl = wb.create_sheet("AVL")
    ws_avl.append(
        [
            "part_number",
            "vendor_code",
            "vendor_part_number",
            "preferred",
            "notes",
            "pb_qty_threshold",
            "pb_unit_cost",
            "pb_lead_days",
        ]
    )
    for part in data.get("parts", []):
        pn = part.get("part_number", "")
        for avl in part.get("avl", []):
            price_breaks = avl.get("price_breaks", [])
            if price_breaks:
                for pb in price_breaks:
                    ws_avl.append(
                        [
                            pn,
                            avl.get("vendor_code", ""),
                            avl.get("vendor_part_number", ""),
                            avl.get("preferred", False),
                            avl.get("notes", ""),
                            pb.get("qty_threshold", ""),
                            pb.get("unit_cost", ""),
                            pb.get("lead_days", ""),
                        ]
                    )
            else:
                # AVL link with no price breaks — emit one row with empty PB columns
                ws_avl.append(
                    [
                        pn,
                        avl.get("vendor_code", ""),
                        avl.get("vendor_part_number", ""),
                        avl.get("preferred", False),
                        avl.get("notes", ""),
                        "",
                        "",
                        "",
                    ]
                )

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# ---------------------------------------------------------------------------
# Import parsing (PLUM-10, D-15/D-16)
# ---------------------------------------------------------------------------


def parse_json_import(content: bytes) -> dict:
    """
    Parse a JSON import file (bytes) into a normalized in-memory dataset dict.

    Expects the schema produced by build_json_export (schema_version=1).
    Returns the parsed dict as-is; validation is performed separately in
    validate_import (two-pass cross-reference check, D-18 Pitfall 6).

    Raises HTTPException 422 on malformed JSON.
    """
    import json as _json

    try:
        data = _json.loads(content)
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=f"Invalid JSON: {exc}",
        )
    if not isinstance(data, dict):
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Import JSON must be an object.",
        )
    return data


def parse_excel_import(content: bytes) -> dict:
    """
    Parse an Excel (.xlsx) import file into the normalized in-memory dataset dict.

    Reads the three sheets (Parts, BOMs, AVL) from the workbook via
    openpyxl.load_workbook(read_only=True, data_only=True) — values only,
    no formula execution (T-06-12 mitigation).

    Returns a dict matching the JSON export schema so validate_import and
    commit_import can handle both formats identically.

    Raises HTTPException 422 on malformed workbook or missing required sheets.
    """
    from io import BytesIO as _BytesIO
    import openpyxl

    try:
        wb = openpyxl.load_workbook(
            _BytesIO(content), read_only=True, data_only=True
        )
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=f"Invalid Excel file: {exc}",
        )

    sheet_names = wb.sheetnames

    # Build parts dict keyed by part_number for assembly
    parts_by_pn: dict[str, dict] = {}

    # --- Parse Parts sheet ---
    if "Parts" in sheet_names:
        ws = wb["Parts"]
        rows = iter(ws.rows)
        next(rows, None)  # skip header
        for row in rows:
            vals = [cell.value for cell in row]
            if not vals or vals[0] is None:
                continue
            pn = str(vals[0]).strip()
            if not pn:
                continue
            tags_raw = vals[4] if len(vals) > 4 and vals[4] is not None else ""
            tags = (
                [t.strip() for t in str(tags_raw).split(",") if t.strip()]
                if tags_raw
                else []
            )
            parts_by_pn[pn] = {
                "part_number": pn,
                "active": True,
                "tags": tags,
                "revisions": [],
                "avl": [],
            }

    # --- Parse BOMs sheet (named "BOM" in export; also accepts "BOMs" for round-trip) ---
    _bom_sheet = next(
        (n for n in ("BOM", "BOMs") if n in sheet_names), None
    )
    if _bom_sheet is not None:
        bom_by_pn_rev: dict[tuple, list] = {}
        ws = wb[_bom_sheet]
        rows = iter(ws.rows)
        next(rows, None)  # skip header
        for row in rows:
            vals = [cell.value for cell in row]
            if not vals or vals[0] is None:
                continue
            parent_pn = str(vals[0]).strip() if vals[0] else ""
            rev_label = str(vals[1]).strip() if len(vals) > 1 and vals[1] else ""
            child_pn = str(vals[2]).strip() if len(vals) > 2 and vals[2] else ""
            qty = (
                str(vals[3]).strip()
                if len(vals) > 3 and vals[3] is not None
                else "1"
            )
            ref_des = vals[4] if len(vals) > 4 else None
            if parent_pn and child_pn:
                key = (parent_pn, rev_label)
                bom_by_pn_rev.setdefault(key, []).append(
                    {
                        "child_part_number": child_pn,
                        "quantity": qty,
                        "reference_designators": (
                            str(ref_des) if ref_des else None
                        ),
                    }
                )
        # Attach BOM lines to parts/revisions
        for (parent_pn, rev_label), lines in bom_by_pn_rev.items():
            if parent_pn not in parts_by_pn:
                parts_by_pn[parent_pn] = {
                    "part_number": parent_pn,
                    "active": True,
                    "tags": [],
                    "revisions": [],
                    "avl": [],
                }
            part_entry = parts_by_pn[parent_pn]
            rev_entry = next(
                (
                    r
                    for r in part_entry["revisions"]
                    if r.get("revision_label") == rev_label
                ),
                None,
            )
            if rev_entry is None:
                rev_entry = {
                    "revision_label": rev_label,
                    "revision_number": 1,
                    "status": "draft",
                    "description": None,
                    "material_cost": None,
                    "sale_price": None,
                    "released_cost_snapshot": None,
                    "selected_vendor_link_id": None,
                    "selected_price_break_index": None,
                    "bom": [],
                }
                part_entry["revisions"].append(rev_entry)
            rev_entry["bom"].extend(lines)

    # --- Parse AVL sheet ---
    if "AVL" in sheet_names:
        ws = wb["AVL"]
        rows = iter(ws.rows)
        next(rows, None)  # skip header
        avl_links_temp: dict[tuple, dict] = {}
        for row in rows:
            vals = [cell.value for cell in row]
            if not vals or vals[0] is None:
                continue
            pn = str(vals[0]).strip() if vals[0] else ""
            vendor_code = (
                str(vals[1]).strip() if len(vals) > 1 and vals[1] else ""
            )
            if not pn or not vendor_code:
                continue
            key = (pn, vendor_code)
            if key not in avl_links_temp:
                avl_links_temp[key] = {
                    "vendor_code": vendor_code,
                    "vendor_part_number": (
                        str(vals[2]).strip()
                        if len(vals) > 2 and vals[2]
                        else None
                    ),
                    "preferred": (
                        bool(vals[3])
                        if len(vals) > 3 and vals[3] is not None
                        else False
                    ),
                    "notes": (
                        str(vals[4]).strip() if len(vals) > 4 and vals[4] else None
                    ),
                    "price_breaks": [],
                }
                if pn not in parts_by_pn:
                    parts_by_pn[pn] = {
                        "part_number": pn,
                        "active": True,
                        "tags": [],
                        "revisions": [],
                        "avl": [],
                    }
                parts_by_pn[pn]["avl"].append(avl_links_temp[key])
            # Price break columns: pb_qty_threshold, pb_unit_cost, pb_lead_days
            pb_qty = vals[5] if len(vals) > 5 and vals[5] is not None else None
            pb_cost = vals[6] if len(vals) > 6 and vals[6] is not None else None
            if pb_qty is not None and pb_cost is not None:
                try:
                    avl_links_temp[key]["price_breaks"].append(
                        {
                            "qty_threshold": int(pb_qty),
                            "unit_cost": str(pb_cost),
                            "lead_days": (
                                int(vals[7])
                                if len(vals) > 7 and vals[7] is not None
                                else None
                            ),
                        }
                    )
                except (ValueError, TypeError):
                    pass  # skip malformed price break row

    return {
        "schema_version": 1,
        "exported_at": None,
        "parts": list(parts_by_pn.values()),
    }


# ---------------------------------------------------------------------------
# Import validate/preview (PLUM-10, D-18 step 1 — NO DB WRITES)
# ---------------------------------------------------------------------------


async def validate_import(
    db: AsyncSession,
    data: dict,
) -> "ImportPreviewResponse":
    """
    Validate an import payload (two-pass cross-reference check, D-18 step 1).

    NEVER writes to the DB — preview is side-effect free.

    Two-pass algorithm (RESEARCH Pitfall 6):
      Pass 1: Collect all part_numbers declared in the file (file-declared set).
      Pass 2: For each row, validate against DB union file-declared parts.
               AVL vendor_code must resolve to an existing syerp_partner (is_vendor=True).

    Counts rows as new vs. updated by checking existence on stable keys:
      - Part: stable key = part_number

    Returns ImportPreviewResponse(new_count, updated_count, errors).
    """
    from app.modules.plum.models import PlumPart
    from app.modules.plum.schemas import ImportPreviewResponse, ImportRowError
    from app.modules.syerp.models import Partner as SyerpPartner

    errors: list[ImportRowError] = []
    parts_in_file: list[dict] = data.get("parts", [])

    # --- Pass 1: Collect file-declared part_numbers ---
    file_part_numbers: set[str] = set()
    for part in parts_in_file:
        pn = part.get("part_number", "")
        if pn:
            file_part_numbers.add(str(pn))

    # --- Fetch existing parts from DB (stable key: part_number) ---
    db_all_parts_result = await db.execute(select(PlumPart.part_number))
    db_all_part_numbers: set[str] = {
        str(row[0]) for row in db_all_parts_result.all()
    }

    # Combined known set for BOM cross-reference validation
    known_part_numbers = db_all_part_numbers | file_part_numbers

    # --- Resolve vendor_codes referenced in the file ---
    file_vendor_codes: set[str] = set()
    for part in parts_in_file:
        for avl in part.get("avl", []):
            vc = avl.get("vendor_code", "") or avl.get("vendor_id", "")
            if vc:
                file_vendor_codes.add(str(vc))

    db_vendor_codes: set[str] = set()
    if file_vendor_codes:
        vend_result = await db.execute(
            select(SyerpPartner.code).where(
                SyerpPartner.is_vendor.is_(True),
                SyerpPartner.code.in_(list(file_vendor_codes)),
            )
        )
        db_vendor_codes = {str(row[0]) for row in vend_result.all()}

    # --- Pass 2: Validate each part + BOM + AVL ---
    new_count = 0
    updated_count = 0

    for row_idx, part in enumerate(parts_in_file, start=1):
        pn = part.get("part_number", "")
        if not pn:
            errors.append(
                ImportRowError(
                    row=row_idx,
                    field="part_number",
                    message="part_number is required.",
                )
            )
            continue

        if pn in db_all_part_numbers:
            updated_count += 1
        else:
            new_count += 1

        # Validate BOM child cross-references
        for rev in part.get("revisions", []):
            for bom_line in rev.get("bom", []):
                child_pn = bom_line.get("child_part_number", "")
                if child_pn and child_pn not in known_part_numbers:
                    errors.append(
                        ImportRowError(
                            row=row_idx,
                            field="child_part_number",
                            message=(
                                f"BOM child part '{child_pn}' is not declared in the "
                                f"import file and does not exist in the database."
                            ),
                        )
                    )

        # Validate AVL vendor_code cross-references
        for avl_idx, avl in enumerate(part.get("avl", []), start=1):
            vc = avl.get("vendor_code", "") or avl.get("vendor_id", "")
            if vc and vc not in db_vendor_codes:
                errors.append(
                    ImportRowError(
                        row=row_idx,
                        field="vendor_id",
                        message=(
                            f"AVL vendor_code '{vc}' does not match any known SYERP "
                            f"vendor (is_vendor=True). Row {avl_idx} for part '{pn}'."
                        ),
                    )
                )

    return ImportPreviewResponse(
        new_count=new_count,
        updated_count=updated_count,
        errors=errors,
    )


# ---------------------------------------------------------------------------
# Import commit (PLUM-10, D-17/D-18 step 2 — upsert, never delete)
# ---------------------------------------------------------------------------


async def commit_import(
    db: AsyncSession,
    data: dict,
    actor_id: str,
) -> "ImportCommitResponse":
    """
    Apply an import payload to the database in one transaction (PLUM-10, D-17/D-18).

    Algorithm (upsert on stable keys, NEVER hard-delete):
      1. Re-validate (same logic as validate_import). Any errors block the commit.
      2. Upsert PlumPart rows by part_number (select-before-insert-or-update).
      3. Upsert PlumPartRevision rows by (part_id, revision_label).
      4. Upsert PlumBomItem rows by (parent_revision_id, child_part_id).
      5. Upsert PlumAvlLink rows by (part_id, vendor_id).
         Price breaks: replace-all per AVL link (subordinate data).
      6. NEVER calls DELETE on parts/revisions/BOM/AVL rows absent from the file. (D-17)
      7. Writes plum.imported audit event.

    Returns ImportCommitResponse(inserted, updated).
    Raises HTTPException 400 if validate_import returns any errors.
    """
    from app.modules.auth.service import write_audit
    from app.modules.plum.models import (
        PlumAvlLink,
        PlumAvlPriceBreak,
        PlumBomItem,
        PlumPart,
        PlumPartRevision,
    )
    from app.modules.plum.schemas import ImportCommitResponse
    from app.modules.syerp.models import Partner as SyerpPartner

    # --- Step 1: Re-validate (D-18 Pitfall 5 — never trust preview result from client) ---
    preview = await validate_import(db, data)
    if preview.errors:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                f"Import validation failed with {len(preview.errors)} error(s). "
                f"Run preview first and resolve all errors before committing."
            ),
        )

    inserted = 0
    updated = 0
    parts_in_file: list[dict] = data.get("parts", [])

    # Fetch vendor_code -> vendor_id mapping for the file's vendor codes
    file_vendor_codes = {
        avl.get("vendor_code", "") or avl.get("vendor_id", "")
        for part in parts_in_file
        for avl in part.get("avl", [])
        if avl.get("vendor_code", "") or avl.get("vendor_id", "")
    }
    vendor_id_by_code: dict[str, str] = {}
    if file_vendor_codes:
        vend_result = await db.execute(
            select(SyerpPartner.id, SyerpPartner.code).where(
                SyerpPartner.is_vendor.is_(True),
                SyerpPartner.code.in_(list(file_vendor_codes)),
            )
        )
        for vid, vcode in vend_result.all():
            vendor_id_by_code[str(vcode)] = str(vid)

    for part_data in parts_in_file:
        pn = str(part_data.get("part_number", "")).strip()
        if not pn:
            continue

        # --- Upsert Part (stable key: part_number) ---
        existing_part_result = await db.execute(
            select(PlumPart).where(PlumPart.part_number == pn)
        )
        existing_part = existing_part_result.scalars().first()
        if existing_part is None:
            new_part = PlumPart(
                id=str(uuid.uuid4()),
                part_number=pn,
                active=bool(part_data.get("active", True)),
            )
            db.add(new_part)
            await db.flush()
            part_id = str(new_part.id)
            inserted += 1
        else:
            part_id = str(existing_part.id)
            if existing_part.active != bool(part_data.get("active", True)):
                existing_part.active = bool(part_data.get("active", True))
            updated += 1

        # --- Upsert Revisions (stable key: part_id + revision_label) ---
        for rev_data in part_data.get("revisions", []):
            rev_label = str(rev_data.get("revision_label", "")).strip()
            if not rev_label:
                continue
            existing_rev_result = await db.execute(
                select(PlumPartRevision).where(
                    PlumPartRevision.part_id == part_id,
                    PlumPartRevision.revision_label == rev_label,
                )
            )
            existing_rev = existing_rev_result.scalars().first()
            if existing_rev is None:
                max_rev_result = await db.execute(
                    select(func.max(PlumPartRevision.revision_number)).where(
                        PlumPartRevision.part_id == part_id
                    )
                )
                max_rev_num = max_rev_result.scalar() or 0
                new_rev = PlumPartRevision(
                    id=str(uuid.uuid4()),
                    part_id=part_id,
                    revision_label=rev_label,
                    revision_number=(
                        rev_data.get("revision_number") or max_rev_num + 1
                    ),
                    status=rev_data.get("status", "draft"),
                    description=rev_data.get("description"),
                    material_cost=(
                        Decimal(str(rev_data["material_cost"]))
                        if rev_data.get("material_cost") is not None
                        else None
                    ),
                    sale_price=(
                        Decimal(str(rev_data["sale_price"]))
                        if rev_data.get("sale_price") is not None
                        else None
                    ),
                    released_cost_snapshot=(
                        Decimal(str(rev_data["released_cost_snapshot"]))
                        if rev_data.get("released_cost_snapshot") is not None
                        else None
                    ),
                    selected_vendor_link_id=rev_data.get("selected_vendor_link_id"),
                    selected_price_break_index=rev_data.get(
                        "selected_price_break_index"
                    ),
                )
                db.add(new_rev)
                await db.flush()
                rev_id = str(new_rev.id)
            else:
                rev_id = str(existing_rev.id)
                existing_rev.status = rev_data.get("status", existing_rev.status)
                existing_rev.description = rev_data.get(
                    "description", existing_rev.description
                )
                if rev_data.get("material_cost") is not None:
                    existing_rev.material_cost = Decimal(
                        str(rev_data["material_cost"])
                    )
                if rev_data.get("sale_price") is not None:
                    existing_rev.sale_price = Decimal(str(rev_data["sale_price"]))

            # --- Upsert BOM lines (stable key: parent_revision_id + child_part_id) ---
            for bom_line in rev_data.get("bom", []):
                child_pn = str(bom_line.get("child_part_number", "")).strip()
                if not child_pn:
                    continue
                child_part_result = await db.execute(
                    select(PlumPart.id).where(PlumPart.part_number == child_pn)
                )
                child_part_id_val = child_part_result.scalar()
                if child_part_id_val is None:
                    continue  # validated in preview; skip if still missing
                child_part_id = str(child_part_id_val)
                existing_bom_result = await db.execute(
                    select(PlumBomItem).where(
                        PlumBomItem.parent_revision_id == rev_id,
                        PlumBomItem.child_part_id == child_part_id,
                    )
                )
                existing_bom = existing_bom_result.scalars().first()
                qty_val = bom_line.get("quantity", "1")
                try:
                    qty_dec = (
                        Decimal(str(qty_val)) if qty_val else Decimal("1")
                    )
                except Exception:
                    qty_dec = Decimal("1")
                if existing_bom is None:
                    max_sort_result = await db.execute(
                        select(func.max(PlumBomItem.sort_order)).where(
                            PlumBomItem.parent_revision_id == rev_id
                        )
                    )
                    next_sort = (max_sort_result.scalar() or 0) + 1
                    db.add(
                        PlumBomItem(
                            id=str(uuid.uuid4()),
                            parent_revision_id=rev_id,
                            child_part_id=child_part_id,
                            qty=qty_dec,
                            ref_des=bom_line.get("reference_designators"),
                            sort_order=next_sort,
                        )
                    )
                else:
                    existing_bom.qty = qty_dec
                    existing_bom.ref_des = bom_line.get("reference_designators")
                await db.flush()

        # --- Upsert AVL links (stable key: part_id + vendor_id) ---
        for avl_data in part_data.get("avl", []):
            vc = str(
                avl_data.get("vendor_code", "") or avl_data.get("vendor_id", "")
            ).strip()
            vendor_db_id = vendor_id_by_code.get(vc)
            if not vendor_db_id:
                continue  # validated in preview; skip if still missing

            existing_avl_result = await db.execute(
                select(PlumAvlLink).where(
                    PlumAvlLink.part_id == part_id,
                    PlumAvlLink.vendor_id == vendor_db_id,
                )
            )
            existing_avl = existing_avl_result.scalars().first()
            if existing_avl is None:
                new_avl = PlumAvlLink(
                    id=str(uuid.uuid4()),
                    part_id=part_id,
                    vendor_id=vendor_db_id,
                    vendor_part_number=avl_data.get("vendor_part_number"),
                    preferred=bool(avl_data.get("preferred", False)),
                    notes=avl_data.get("notes"),
                    active=True,
                )
                db.add(new_avl)
                await db.flush()
                avl_link_id = str(new_avl.id)
            else:
                avl_link_id = str(existing_avl.id)
                existing_avl.vendor_part_number = avl_data.get(
                    "vendor_part_number"
                )
                existing_avl.preferred = bool(avl_data.get("preferred", False))
                existing_avl.notes = avl_data.get("notes")

            # Replace-all price breaks for this AVL link.
            # NEVER-DELETE rule applies to parts/revisions/BOM/AVL — not to
            # subordinate price-break rows (which are replace-all per link, D-17).
            existing_pbs_result = await db.execute(
                select(PlumAvlPriceBreak).where(
                    PlumAvlPriceBreak.avl_link_id == avl_link_id
                )
            )
            for old_pb in existing_pbs_result.scalars().all():
                await db.delete(old_pb)
            await db.flush()

            for pb_idx, pb_data in enumerate(avl_data.get("price_breaks", [])):
                try:
                    unit_cost_dec = Decimal(
                        str(pb_data.get("unit_cost", "0"))
                    )
                except Exception:
                    unit_cost_dec = Decimal("0")
                db.add(
                    PlumAvlPriceBreak(
                        id=str(uuid.uuid4()),
                        avl_link_id=avl_link_id,
                        qty_threshold=int(pb_data.get("qty_threshold", 1)),
                        unit_cost=unit_cost_dec,
                        lead_days=pb_data.get("lead_days"),
                        sort_order=pb_idx,
                    )
                )
            await db.flush()

    # NEVER hard-delete parts, revisions, BOM lines, or AVL links absent from the file. (D-17)
    # Only price breaks are replaced (replace-all per AVL link, subordinate data).

    await db.commit()

    await write_audit(
        db,
        actor_id=actor_id,
        action="plum.imported",
        target_type="import",
        target_id="",
        detail=f"Import committed: {inserted} inserted, {updated} updated.",
    )

    return ImportCommitResponse(inserted=inserted, updated=updated)
